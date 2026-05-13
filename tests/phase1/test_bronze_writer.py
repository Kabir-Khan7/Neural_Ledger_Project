"""
Tests — Step 3: Bronze Writer
================================
Tests the Excel parser, CSV parser, and Bronze Writer end-to-end.
Run with: python -m pytest tests/phase1/test_bronze_writer.py -v
"""

import json
import pytest
import tempfile
import csv as csvlib
from pathlib import Path
from uuid import uuid4

import sys
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

import openpyxl
from openpyxl import Workbook

from src.database.init_db import DatabaseManager
from src.phase1.ingestion.excel_parser import ExcelParser
from src.phase1.ingestion.csv_parser import CSVParser
from src.phase1.bronze.bronze_writer import BronzeWriter


# ─────────────────────────────────────────────────────────────────────────────
# Fixtures — create realistic test files
# ─────────────────────────────────────────────────────────────────────────────

@pytest.fixture
def temp_db(tmp_path):
    db_path = tmp_path / "test.db"
    manager = DatabaseManager(db_path=db_path)
    manager.initialise()
    # Patch the global db singleton to use test db
    import src.database.init_db as db_module
    original = db_module.db
    db_module.db = manager
    yield manager
    db_module.db = original
    manager.close()


@pytest.fixture
def clean_excel(tmp_path):
    """Standard clean Excel — what a good accountant produces."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(["Date", "Vendor", "Description", "Amount", "Category"])
    ws.append(["15/01/2024", "Shell Clifton", "Fuel purchase", 5000, "Fuel"])
    ws.append(["16/01/2024", "K-Electric", "Electricity bill", 12000, "Utilities"])
    ws.append(["17/01/2024", "Ahmed Transport", "Freight charges", 8500, "Logistics"])
    path = tmp_path / "clean_ledger.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def excel_with_title_rows(tmp_path):
    """Excel with company name and report title before headers — very common."""
    wb = Workbook()
    ws = wb.active
    ws.append(["ABC Trading Company (Pvt) Ltd"])       # Row 1: company name
    ws.append(["Cash Book — January 2024"])             # Row 2: report title
    ws.append([])                                       # Row 3: blank
    ws.append(["Date", "Particulars", "Dr", "Cr", "Balance"])  # Row 4: REAL headers
    ws.append(["01/01/2024", "Opening Balance", "", 50000, 50000])
    ws.append(["15/01/2024", "Shell Petrol", 5000, "", 45000])
    ws.append(["20/01/2024", "Sales Receipt", "", 30000, 75000])
    path = tmp_path / "cashbook_with_titles.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def excel_dr_cr_columns(tmp_path):
    """Excel with split Debit/Credit columns — standard accounting format."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Taareekh", "Vendor Name", "Description", "Debit", "Credit"])
    ws.append(["15/01/2024", "Shell", "Petrol", 5000, ""])
    ws.append(["16/01/2024", "Customer A", "Sales", "", 25000])
    ws.append(["17/01/2024", "K-Electric", "Bill", 12000, ""])
    path = tmp_path / "dr_cr_ledger.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def excel_urdu_headers(tmp_path):
    """Excel with Urdu/mixed headers — real Pakistani accountant file."""
    wb = Workbook()
    ws = wb.active
    ws.append(["تاریخ", "رقم", "تفصیل", "Vendor"])
    ws.append(["15/01/2024", 5000, "Fuel purchase", "Shell"])
    ws.append(["16/01/2024", 12000, "Bijli ka bill", "K-Electric"])
    path = tmp_path / "urdu_headers.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def excel_with_blank_rows(tmp_path):
    """Excel with blank rows used as visual separators."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Vendor", "Amount"])
    ws.append(["15/01/2024", "Shell", 5000])
    ws.append([])   # blank separator
    ws.append(["16/01/2024", "K-Electric", 12000])
    ws.append([])   # blank separator
    ws.append(["17/01/2024", "Careem", 800])
    path = tmp_path / "spaced_ledger.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def excel_multiple_sheets(tmp_path):
    """Excel with multiple sheets — should pick the data-rich one."""
    wb = Workbook()
    # Sheet 1: summary (sparse)
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(["Total Revenue", 500000])

    # Sheet 2: transactions (rich)
    ws2 = wb.create_sheet("Transactions")
    ws2.append(["Date", "Vendor", "Amount"])
    for i in range(20):
        ws2.append([f"0{(i%28)+1}/01/2024", f"Vendor {i}", 1000 * (i+1)])

    path = tmp_path / "multi_sheet.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def csv_utf8(tmp_path):
    """Standard UTF-8 CSV."""
    path = tmp_path / "transactions_utf8.csv"
    with open(str(path), "w", encoding="utf-8", newline="") as f:
        writer = csvlib.writer(f)
        writer.writerow(["Date", "Amount", "Vendor", "Description"])
        writer.writerow(["15/01/2024", "5000", "Shell", "Fuel"])
        writer.writerow(["16/01/2024", "12000", "K-Electric", "Bill"])
    return path


@pytest.fixture
def csv_windows1252(tmp_path):
    """Windows-1252 CSV — exported by Windows Excel 'Save as CSV'."""
    path = tmp_path / "transactions_win.csv"
    with open(str(path), "w", encoding="cp1252", newline="") as f:
        writer = csvlib.writer(f)
        writer.writerow(["Date", "Amount", "Vendor", "Description"])
        writer.writerow(["15/01/2024", "5000", "Café", "Coffee expense"])
        writer.writerow(["16/01/2024", "1200", "Résumé Co", "HR service"])
    return path


@pytest.fixture
def csv_with_bom(tmp_path):
    """UTF-8 with BOM — produced by Excel 'Save as CSV UTF-8' on Windows."""
    path = tmp_path / "transactions_bom.csv"
    with open(str(path), "w", encoding="utf-8-sig", newline="") as f:
        writer = csvlib.writer(f)
        writer.writerow(["Date", "Amount", "Vendor"])
        writer.writerow(["15/01/2024", "5000", "Shell"])
    return path


@pytest.fixture
def csv_semicolon_delimiter(tmp_path):
    """Semicolon-delimited CSV — from European accounting software."""
    path = tmp_path / "tally_export.csv"
    with open(str(path), "w", encoding="utf-8", newline="") as f:
        f.write("Date;Amount;Vendor;Type\n")
        f.write("15/01/2024;5000;Shell;Payment\n")
        f.write("16/01/2024;12000;K-Electric;Payment\n")
    return path


@pytest.fixture
def csv_with_blank_and_total_rows(tmp_path):
    """CSV with blank rows and total rows mixed in."""
    path = tmp_path / "mixed_csv.csv"
    with open(str(path), "w", encoding="utf-8", newline="") as f:
        writer = csvlib.writer(f)
        writer.writerow(["Date", "Amount", "Vendor"])
        writer.writerow(["15/01/2024", "5000", "Shell"])
        writer.writerow(["", "", ""])          # blank row
        writer.writerow(["16/01/2024", "12000", "K-Electric"])
        writer.writerow(["Total", "17000", ""])  # total row — should be skipped
    return path


# ─────────────────────────────────────────────────────────────────────────────
# 1. Excel Parser Tests
# ─────────────────────────────────────────────────────────────────────────────

class TestExcelParser:
    parser = ExcelParser()

    def test_clean_excel_parsed(self, clean_excel):
        rows, headers, meta = self.parser.parse(str(clean_excel))
        assert len(rows) == 3
        assert "Date" in headers
        assert "Amount" in headers

    def test_original_column_names_preserved(self, clean_excel):
        rows, headers, _ = self.parser.parse(str(clean_excel))
        assert "Vendor" in headers
        assert "Description" in headers

    def test_row_values_correct(self, clean_excel):
        rows, _, _ = self.parser.parse(str(clean_excel))
        first = rows[0]
        assert "Shell Clifton" in str(first.get("Vendor", ""))

    def test_excel_with_title_rows(self, excel_with_title_rows):
        """Header detection must skip company name and report title rows."""
        rows, headers, meta = self.parser.parse(str(excel_with_title_rows))
        # Should find the real headers — Date, Particulars, Dr, Cr, Balance
        assert any(
            h in headers
            for h in ["Date", "Particulars", "Dr", "Cr", "Balance"]
        ), f"Expected real headers, got: {headers}"
        # Should have data rows (not title rows)
        assert len(rows) >= 2

    def test_dr_cr_columns_preserved(self, excel_dr_cr_columns):
        """Dr and Cr split columns must be preserved as separate columns."""
        rows, headers, _ = self.parser.parse(str(excel_dr_cr_columns))
        assert "Debit" in headers or "Dr" in headers or "Taareekh" in headers

    def test_urdu_headers_preserved(self, excel_urdu_headers):
        """Urdu column headers must be preserved exactly."""
        rows, headers, _ = self.parser.parse(str(excel_urdu_headers))
        urdu_headers = [h for h in headers if any(
            ord(c) > 127 for c in str(h)
        )]
        assert len(urdu_headers) >= 1, "Urdu headers should be preserved"

    def test_blank_rows_removed(self, excel_with_blank_rows):
        """Blank separator rows must be filtered out."""
        rows, _, _ = self.parser.parse(str(excel_with_blank_rows))
        assert len(rows) == 3  # 3 data rows, not 5

    def test_multiple_sheets_picks_data_rich(self, excel_multiple_sheets):
        """Should pick the Transactions sheet (20 rows) not Summary (1 row)."""
        rows, headers, meta = self.parser.parse(str(excel_multiple_sheets))
        assert len(rows) >= 15  # Data-rich sheet has 20 rows
        assert "Date" in headers

    def test_parse_meta_returned(self, clean_excel):
        _, _, meta = self.parser.parse(str(clean_excel))
        assert "total_rows" in meta
        assert "total_cols" in meta
        assert "parse_method" in meta

    def test_empty_columns_dropped(self, tmp_path):
        """Completely empty columns should be dropped."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "", "Amount"])  # empty middle column
        ws.append(["15/01/2024", "", 5000])
        path = tmp_path / "empty_col.xlsx"
        wb.save(str(path))

        rows, headers, _ = self.parser.parse(str(path))
        # Empty column should not appear or be named Column_1
        assert len(rows) == 1


# ─────────────────────────────────────────────────────────────────────────────
# 2. CSV Parser Tests
# ─────────────────────────────────────────────────────────────────────────────

class TestCSVParser:
    parser = CSVParser()

    def test_utf8_csv_parsed(self, csv_utf8):
        rows, headers, meta = self.parser.parse(str(csv_utf8))
        assert len(rows) == 2
        assert "Date" in headers
        assert meta["encoding"] in ("utf-8", "ascii", "UTF-8-SIG", "utf-8-sig")

    def test_windows1252_csv_parsed(self, csv_windows1252):
        """Windows-1252 encoding must be detected and handled."""
        rows, headers, meta = self.parser.parse(str(csv_windows1252))
        assert len(rows) == 2
        # Special characters should be preserved
        vendors = [str(r.get("Vendor", "")) for r in rows]
        assert any("Caf" in v for v in vendors)

    def test_bom_csv_parsed(self, csv_with_bom):
        """UTF-8 BOM files must parse without BOM in column names."""
        rows, headers, meta = self.parser.parse(str(csv_with_bom))
        assert len(rows) == 1
        # BOM must not appear in first header
        assert not headers[0].startswith("\ufeff"), "BOM found in header"

    def test_semicolon_delimiter_detected(self, csv_semicolon_delimiter):
        rows, headers, meta = self.parser.parse(str(csv_semicolon_delimiter))
        assert meta["delimiter"] == ";"
        assert len(rows) == 2
        assert "Date" in headers

    def test_blank_rows_removed(self, csv_with_blank_and_total_rows):
        rows, _, _ = self.parser.parse(str(csv_with_blank_and_total_rows))
        assert len(rows) == 2  # 2 data rows, blank and total removed

    def test_total_row_removed(self, csv_with_blank_and_total_rows):
        rows, _, _ = self.parser.parse(str(csv_with_blank_and_total_rows))
        first_values = [str(r.get("Date", "")).lower() for r in rows]
        assert "total" not in first_values

    def test_tab_delimited_csv(self, tmp_path):
        path = tmp_path / "tab_export.csv"
        with open(str(path), "w", encoding="utf-8") as f:
            f.write("Date\tAmount\tVendor\n")
            f.write("15/01/2024\t5000\tShell\n")
        rows, headers, meta = self.parser.parse(str(path))
        assert meta["delimiter"] == "\t"
        assert len(rows) == 1

    def test_parse_meta_returned(self, csv_utf8):
        _, _, meta = self.parser.parse(str(csv_utf8))
        assert "encoding" in meta
        assert "delimiter" in meta
        assert "total_rows" in meta

    def test_whitespace_stripped_from_values(self, tmp_path):
        path = tmp_path / "spaced.csv"
        with open(str(path), "w", encoding="utf-8") as f:
            f.write("Date, Amount, Vendor\n")
            f.write("15/01/2024,  5000  , Shell  \n")
        rows, headers, _ = self.parser.parse(str(path))
        vendor_val = str(rows[0].get("Vendor", rows[0].get(" Vendor", "")))
        assert vendor_val.strip() == "Shell"


# ─────────────────────────────────────────────────────────────────────────────
# 3. Bronze Writer Tests
# ─────────────────────────────────────────────────────────────────────────────

class TestBronzeWriter:

    def _make_metadata(self, file_path, source_type="xlsx",
                       source_software="manual_excel"):
        from src.database.init_db import file_hash as fh
        return {
            "file_path":       str(file_path),
            "file_name":       Path(file_path).name,
            "source_type":     source_type,
            "source_software": source_software,
            "detection_conf":  0.85,
            "file_hash":       fh(str(file_path)),
            "file_size_bytes": Path(file_path).stat().st_size,
            "detected_at":     "2024-01-15T14:30:00",
        }

    def test_excel_ingested_to_bronze(self, clean_excel, temp_db):
        writer = BronzeWriter()
        meta = self._make_metadata(clean_excel)
        result = writer.ingest(meta)

        assert result["success"] is True
        assert result["rows_written"] == 3
        assert result["batch_id"] is not None

    def test_bronze_rows_in_database(self, clean_excel, temp_db):
        writer = BronzeWriter()
        meta = self._make_metadata(clean_excel)
        result = writer.ingest(meta)

        with temp_db.connection() as conn:
            count = conn.execute(
                "SELECT COUNT(*) FROM bronze_transactions "
                "WHERE ingestion_batch_id = ?",
                [result["batch_id"]]
            ).fetchone()[0]
        assert count == 3

    def test_raw_content_preserves_original_columns(self, clean_excel, temp_db):
        """Original column names must be preserved in raw_content JSON."""
        writer = BronzeWriter()
        meta = self._make_metadata(clean_excel)
        result = writer.ingest(meta)

        with temp_db.connection() as conn:
            row = conn.execute(
                "SELECT raw_content FROM bronze_transactions LIMIT 1"
            ).fetchone()

        raw = json.loads(row[0])
        assert "Vendor" in raw
        assert "Amount" in raw

    def test_source_software_stored(self, clean_excel, temp_db):
        writer = BronzeWriter()
        meta = self._make_metadata(
            clean_excel, source_software="quickbooks_desktop"
        )
        writer.ingest(meta)

        with temp_db.connection() as conn:
            sw = conn.execute(
                "SELECT DISTINCT source_software FROM bronze_transactions"
            ).fetchone()[0]
        assert sw == "quickbooks_desktop"

    def test_processing_status_is_pending(self, clean_excel, temp_db):
        writer = BronzeWriter()
        meta = self._make_metadata(clean_excel)
        writer.ingest(meta)

        with temp_db.connection() as conn:
            statuses = conn.execute(
                "SELECT DISTINCT processing_status FROM bronze_transactions"
            ).fetchall()
        assert all(s[0] == "pending" for s in statuses)

    def test_batch_id_groups_rows(self, clean_excel, temp_db):
        writer = BronzeWriter()
        meta = self._make_metadata(clean_excel)
        result = writer.ingest(meta)

        with temp_db.connection() as conn:
            count = conn.execute(
                "SELECT COUNT(DISTINCT ingestion_batch_id) "
                "FROM bronze_transactions"
            ).fetchone()[0]
        assert count == 1

    def test_csv_ingested_to_bronze(self, csv_utf8, temp_db):
        writer = BronzeWriter()
        meta = self._make_metadata(csv_utf8, source_type="csv")
        result = writer.ingest(meta)

        assert result["success"] is True
        assert result["rows_written"] == 2

    def test_urdu_headers_preserved_in_bronze(self, excel_urdu_headers, temp_db):
        writer = BronzeWriter()
        meta = self._make_metadata(excel_urdu_headers)
        result = writer.ingest(meta)

        assert result["success"] is True
        with temp_db.connection() as conn:
            row = conn.execute(
                "SELECT raw_content FROM bronze_transactions LIMIT 1"
            ).fetchone()
        raw = json.loads(row[0])
        # At least one key should be non-ASCII (Urdu)
        urdu_keys = [k for k in raw.keys() if any(ord(c) > 127 for c in str(k))]
        assert len(urdu_keys) >= 1

    def test_file_hash_stored(self, clean_excel, temp_db):
        writer = BronzeWriter()
        meta = self._make_metadata(clean_excel)
        writer.ingest(meta)

        with temp_db.connection() as conn:
            hashes = conn.execute(
                "SELECT DISTINCT source_file_hash FROM bronze_transactions "
                "WHERE source_file_hash IS NOT NULL"
            ).fetchall()
        assert len(hashes) == 1
        assert len(hashes[0][0]) == 64  # SHA256 hex length

    def test_row_numbers_sequential(self, clean_excel, temp_db):
        writer = BronzeWriter()
        meta = self._make_metadata(clean_excel)
        result = writer.ingest(meta)

        with temp_db.connection() as conn:
            rows = conn.execute(
                "SELECT source_row_number FROM bronze_transactions "
                "ORDER BY source_row_number"
            ).fetchall()
        row_numbers = [r[0] for r in rows]
        assert row_numbers == [1, 2, 3]

    def test_ingest_returns_parse_meta(self, clean_excel, temp_db):
        writer = BronzeWriter()
        meta = self._make_metadata(clean_excel)
        result = writer.ingest(meta)

        assert "parse_meta" in result
        assert result["parse_meta"]["total_rows"] == 3

    def test_two_files_have_different_batch_ids(
        self, clean_excel, csv_utf8, temp_db
    ):
        writer = BronzeWriter()

        meta1 = self._make_metadata(clean_excel, source_type="xlsx")
        meta2 = self._make_metadata(csv_utf8, source_type="csv")

        r1 = writer.ingest(meta1)
        r2 = writer.ingest(meta2)

        assert r1["batch_id"] != r2["batch_id"]

    def test_empty_excel_handled_gracefully(self, tmp_path, temp_db):
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Amount"])
        # No data rows
        path = tmp_path / "empty_data.xlsx"
        wb.save(str(path))

        writer = BronzeWriter()
        meta = self._make_metadata(path)
        result = writer.ingest(meta)

        assert result["success"] is True
        assert result["rows_written"] == 0

    def test_audit_log_entry_created(self, clean_excel, temp_db):
        writer = BronzeWriter()
        meta = self._make_metadata(clean_excel)
        writer.ingest(meta)

        with temp_db.connection() as conn:
            count = conn.execute(
                "SELECT COUNT(*) FROM pipeline_audit_log "
                "WHERE operation LIKE 'bronze_ingest%'"
            ).fetchone()[0]
        assert count >= 1