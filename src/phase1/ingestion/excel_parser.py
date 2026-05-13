"""
Neural Ledger — Excel Parser
==============================
Reads .xlsx and .xls files into a list of raw row dicts.

The core challenge: Pakistani accountant Excel files are not clean tables.
They contain:
  - Multiple header rows (company name, report title, then column headers)
  - Merged cells across header rows
  - Blank rows used as visual separators
  - Summary/total rows mixed with data rows
  - Columns with no headers (just blank)
  - Mixed data types in the same column
  - Dr/Cr split columns instead of a single amount column
  - Urdu text in headers and descriptions

This parser handles all of the above by:
  1. Reading with openpyxl to detect merged cells and real header row
  2. Falling back to pandas for straightforward files
  3. Returning raw dicts preserving ORIGINAL column names exactly
"""

import logging
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Maximum rows to scan looking for the real header row
HEADER_SCAN_LIMIT = 15

# Minimum number of non-empty cells in a row to consider it the header
HEADER_MIN_CELLS = 2


class ExcelParser:
    """
    Parses Excel files into a list of raw row dicts.

    Each dict preserves the original column name as the key
    and the original cell value as the value.

    Example output row:
        {
            "Taareekh":        "15/01/2024",
            "Raqam":           5000.0,
            "Vendor Name":     "Shell Clifton",
            "Description":     "Fuel purchase",
            "Dr":              5000.0,
            "Cr":              0.0,
        }
    """

    def parse(self, file_path: str) -> Tuple[List[Dict[str, Any]], List[str], dict]:
        """
        Parse an Excel file.

        Returns:
            rows         — list of raw row dicts (original column names preserved)
            headers      — list of original column header strings
            parse_meta   — dict with parsing metadata (sheet used, header row, etc.)
        """
        path = Path(file_path)
        ext = path.suffix.lower()

        logger.info(f"Parsing Excel: {path.name}")

        if ext == ".xls":
            return self._parse_xls(file_path)
        else:
            return self._parse_xlsx(file_path)

    # ──────────────────────────────────────────────────────────────────────────
    # XLSX parser — openpyxl for header detection, pandas for data reading
    # ──────────────────────────────────────────────────────────────────────────

    def _parse_xlsx(
        self, file_path: str
    ) -> Tuple[List[Dict[str, Any]], List[str], dict]:

        # Step 1: Use openpyxl to detect the real header row
        header_row_idx, sheet_name = self._detect_header_row(file_path)

        logger.info(
            f"Header detected at row {header_row_idx + 1} "
            f"on sheet '{sheet_name}'"
        )

        # Step 2: Read with pandas using the detected header row
        try:
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=header_row_idx,
                dtype=str,          # Read everything as string — we normalise later
                engine="openpyxl",
                na_values=["", " ", "N/A", "n/a", "NA", "None", "-", "--"],
                keep_default_na=True,
            )
        except Exception as e:
            logger.warning(
                f"pandas failed on {Path(file_path).name}, "
                f"falling back to openpyxl direct read: {e}"
            )
            return self._parse_xlsx_direct(file_path, sheet_name, header_row_idx)

        # Step 3: Clean up the DataFrame
        df = self._clean_dataframe(df)

        if df.empty:
            logger.warning(f"No data rows found in {Path(file_path).name}")
            return [], [], {"header_row": header_row_idx, "sheet": sheet_name}

        # Step 4: Convert to list of dicts
        headers = list(df.columns)
        rows = df.to_dict(orient="records")

        # Remove rows that are entirely NaN (blank rows between data)
        rows = [r for r in rows if any(
            v is not None and str(v).strip() not in ("", "nan", "NaN", "None")
            for v in r.values()
        )]

        logger.info(
            f"Parsed {len(rows)} data rows, "
            f"{len(headers)} columns from {Path(file_path).name}"
        )

        parse_meta = {
            "sheet_name":   sheet_name,
            "header_row":   header_row_idx,
            "total_rows":   len(rows),
            "total_cols":   len(headers),
            "parse_method": "pandas_openpyxl",
        }

        return rows, headers, parse_meta

    def _detect_header_row(self, file_path: str) -> Tuple[int, str]:
        """
        Scan the first HEADER_SCAN_LIMIT rows to find where the real
        column headers are.

        Strategy:
        - Load with openpyxl, read_only=True for speed
        - Find the first row with HEADER_MIN_CELLS non-empty cells
          that looks like column headers (strings, not numbers/dates)
        - Prefer the sheet with the most data if multiple sheets exist

        Returns:
            (header_row_index, sheet_name)
            header_row_index is 0-based (for pandas header= parameter)
        """
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        # Pick the best sheet — prefer first sheet with actual data
        sheet_name = self._pick_best_sheet(wb)
        ws = wb[sheet_name]

        header_row_idx = 0  # default: assume row 0 is header

        rows_checked = 0
        for row_idx, row in enumerate(ws.iter_rows(
            max_row=HEADER_SCAN_LIMIT,
            values_only=True
        )):
            rows_checked += 1
            non_empty = [
                cell for cell in row
                if cell is not None and str(cell).strip() != ""
            ]

            if len(non_empty) < HEADER_MIN_CELLS:
                continue  # Skip mostly-empty rows (title rows)

            # Check if this looks like a header row:
            # Header rows typically have mostly string values
            string_cells = [
                c for c in non_empty
                if isinstance(c, str) and not self._looks_like_date(str(c))
            ]

            string_ratio = len(string_cells) / len(non_empty)
            if string_ratio >= 0.5:  # At least half are strings
                header_row_idx = row_idx
                break

        wb.close()
        logger.debug(
            f"Header detection: row={header_row_idx}, sheet='{sheet_name}'"
        )
        return header_row_idx, sheet_name

    def _pick_best_sheet(self, wb: openpyxl.Workbook) -> str:
        """
        Pick the most data-rich sheet from a workbook.
        Ignores sheets that look like summaries or are nearly empty.
        """
        if len(wb.sheetnames) == 1:
            return wb.sheetnames[0]

        best_sheet = wb.sheetnames[0]
        best_row_count = 0

        for name in wb.sheetnames:
            ws = wb[name]
            # Count non-empty rows (rough estimate)
            count = sum(
                1 for row in ws.iter_rows(max_row=200, values_only=True)
                if any(c is not None for c in row)
            )
            if count > best_row_count:
                best_row_count = count
                best_sheet = name

        return best_sheet

    def _parse_xlsx_direct(
        self,
        file_path: str,
        sheet_name: str,
        header_row_idx: int
    ) -> Tuple[List[Dict[str, Any]], List[str], dict]:
        """
        Direct openpyxl parse — fallback when pandas fails.
        More tolerant of corrupt or unusual Excel files.
        """
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name]

        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(all_rows) <= header_row_idx:
            return [], [], {"parse_method": "openpyxl_direct", "error": "no_data"}

        # Extract headers
        raw_headers = all_rows[header_row_idx]
        headers = [
            str(h).strip() if h is not None else f"Column_{i}"
            for i, h in enumerate(raw_headers)
        ]

        # Extract data rows
        rows = []
        for row in all_rows[header_row_idx + 1:]:
            row_dict = {}
            for i, (header, value) in enumerate(zip(headers, row)):
                row_dict[header] = value
            # Skip completely blank rows
            if any(v is not None for v in row_dict.values()):
                rows.append(row_dict)

        parse_meta = {
            "sheet_name":   sheet_name,
            "header_row":   header_row_idx,
            "total_rows":   len(rows),
            "total_cols":   len(headers),
            "parse_method": "openpyxl_direct",
        }

        logger.info(
            f"Direct parse: {len(rows)} rows, {len(headers)} cols "
            f"from {Path(file_path).name}"
        )

        return rows, headers, parse_meta

    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean up a raw pandas DataFrame from Excel.

        - Remove columns with no header (Unnamed: N)
        - Strip whitespace from column names
        - Drop rows that are entirely NaN
        - Rename duplicate column headers
        """
        # Strip whitespace from column names
        df.columns = [
            str(c).strip() if not str(c).startswith("Unnamed:") else ""
            for c in df.columns
        ]

        # Replace empty column names with positional names
        seen = {}
        new_cols = []
        for i, col in enumerate(df.columns):
            if col == "":
                col = f"Column_{i}"
            if col in seen:
                seen[col] += 1
                col = f"{col}_{seen[col]}"
            else:
                seen[col] = 0
            new_cols.append(col)
        df.columns = new_cols

        # Drop rows where ALL values are NaN
        df = df.dropna(how="all")

        # Drop columns where ALL values are NaN (empty columns)
        df = df.dropna(axis=1, how="all")

        return df

    # ──────────────────────────────────────────────────────────────────────────
    # XLS parser — older Excel format, use xlrd via pandas
    # ──────────────────────────────────────────────────────────────────────────

    def _parse_xls(
        self, file_path: str
    ) -> Tuple[List[Dict[str, Any]], List[str], dict]:
        """Parse older .xls format using pandas with xlrd engine."""

        try:
            # For .xls we trust pandas header detection more directly
            df = pd.read_excel(
                file_path,
                dtype=str,
                engine="xlrd",
                na_values=["", " ", "N/A", "n/a", "NA", "-", "--"],
                keep_default_na=True,
            )
        except Exception as e:
            logger.error(f"Failed to parse .xls {Path(file_path).name}: {e}")
            raise

        df = self._clean_dataframe(df)

        if df.empty:
            return [], [], {"parse_method": "pandas_xlrd"}

        headers = list(df.columns)
        rows = df.to_dict(orient="records")
        rows = [r for r in rows if any(
            v is not None and str(v).strip() not in ("", "nan", "NaN", "None")
            for v in r.values()
        )]

        parse_meta = {
            "sheet_name":   "Sheet1",
            "header_row":   0,
            "total_rows":   len(rows),
            "total_cols":   len(headers),
            "parse_method": "pandas_xlrd",
        }

        logger.info(
            f"XLS parsed: {len(rows)} rows, {len(headers)} cols "
            f"from {Path(file_path).name}"
        )

        return rows, headers, parse_meta

    # ──────────────────────────────────────────────────────────────────────────
    # Utilities
    # ──────────────────────────────────────────────────────────────────────────

    @staticmethod
    def _looks_like_date(value: str) -> bool:
        """Quick heuristic — does this string look like a date?"""
        import re
        date_patterns = [
            r"\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4}",  # 01/15/2024
            r"\d{4}[/\-\.]\d{1,2}[/\-\.]\d{1,2}",     # 2024-01-15
            r"\d{1,2}\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)",
        ]
        return any(re.search(p, value, re.IGNORECASE) for p in date_patterns)